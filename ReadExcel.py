
#Import Package
import openpyxl


#Load Workbook

wk = openpyxl.load_workbook(r"C:\Users\shill\PycharmProjects\Excel\TestPyExcel.xlsx")

sheets = wk.sheetnames
print('-------------------')
print(sheets[4])
print('-------------------')
print("Active Sheet " + wk.active.title)
print('-------------------')
sh = wk[sheets[1]]
print('-------------------')
print(sh.title)
print('-------------------')
print(sh['A3'].value)
print(sh['B4'].value)

print('-------------------')
c1 = sh.cell(3, 2)
print(c1.value)

print('-------------------')
print(c1.row)
print(c1.column)

#fine rows having data

rows = sh.max_row
columns = sh.max_column

print('-------------------')

print("total rows are - " + str(rows))
print("total columns are - " + str(columns))


#for i in range(1, rows+1):
 #   for j in range(1, columns+1):
  #      c =sh.cell(i, j)
   #     print(c.value)

print('-------------------')
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)



