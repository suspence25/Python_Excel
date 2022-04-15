
#Import Package
import openpyxl


#Load Workbook

wk = openpyxl.load_workbook(r"C:\Users\shill\PycharmProjects\Excel\TestCases.xlsx")


sheets = wk.sheetnames
sh = wk[sheets[0]]


first_row = sh.min_row
last_rows = sh.max_row
max_columns = sh.max_column

print('first row: {} last row:{} max columns:{}'.format(first_row,last_rows,max_columns))



print('_____________________________')

for i in range(first_row, last_rows + 1):
    #print(i)
    output = sh['A'+str(i)].value
    print(str(i) + '. ' + str(output))

print('_____________________________')
