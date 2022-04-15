#Import Package
import openpyxl


#Load Workbook

#wk = openpyxl.load_workbook(r"C:\Users\shill\PycharmProjects\Excel\TestPyExcel.xlsx")
wk = openpyxl.Workbook()


sh = wk.active
sh.title = "HelloTestingWorld"

print('-------------------')
print('-------------------')

print(sh.title.value)


sh['A4'].value = 'www.google.com'

#2nd sheet is created

wk.create_sheet(title="Google")

sh1 = wk['Google']
sh1['A3'] = 'gmail'
sh1['A4'] = 'Google Drive'


wk.remove(wk['HelloTestingWorld'])



wk.create_sheet(title="Yahoo", index=1)

wk.save(r"C:\Users\shill\PycharmProjects\Excel\TestWritePyExcel.xlsx")
