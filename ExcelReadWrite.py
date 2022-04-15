import openpyxl

wk = openpyxl.load_workbook(r"C:\Users\shill\PycharmProjects\Excel\ReadWrite.xlsx")

#read

sheets = wk.sheetnames
print(sheets[0])
print(sheets[1])

sh1 = wk[sheets[0]]
print(sh1['A3'].value)

#write


sh2 = wk[sheets[0]]
A6 = sh2.cell(row=6, column=1)

A6.value = "Write Please"

print(sh2['A4'].value)

wk.save(r"C:\Users\shill\PycharmProjects\Excel\ReadWrite.xlsx")